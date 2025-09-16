# ingest/management/commands/import_excel.py
from __future__ import annotations

import os
import re
import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from django.apps import apps

from openpyxl import load_workbook
from dateutil import parser as dtparser

from ingest.models import (
    Workbook, Sheet, ImportBatch,
    Dataset, DatasetRow,
)
from ingest.utils import excel_templates as xt


# ---------- утилиты ----------

KEYVAL_RE = re.compile(r"(?P<key>[A-Za-zА-Яа-я0-9_#\-]+)\s*[:=]\s*(?P<val>[^;,\]\)]+)")

def norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()

def json_sanitize(v: Any) -> Any:
    if isinstance(v, Decimal):
        try:
            return float(v)
        except Exception:
            return str(v)
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return v

def coerce_number(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "").replace("\u00A0", "").replace(",", ".")
    if s == "":
        return None
    try:
        return Decimal(s)
    except Exception:
        return None

def coerce_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    try:
        return dtparser.parse(str(v).strip(), dayfirst=True).date()
    except Exception:
        return None

def sha256_of_file(path: str, chunk: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk)
            if not b: break
            h.update(b)
    return h.hexdigest()

def detect_header_row(ws, max_probe: int = 25) -> int:
    best_row, best_score = 1, -1
    R = min(ws.max_row, max_probe)
    C = ws.max_column
    for r in range(1, R + 1):
        values = [str(ws.cell(row=r, column=c).value or "").strip() for c in range(1, C + 1)]
        nonempty = [v for v in values if v]
        textish = sum(1 for v in nonempty if not v.replace(".", "", 1).isdigit())
        uniq = len(set(v.lower() for v in nonempty))
        score = textish + uniq * 0.5
        if score > best_score and nonempty:
            best_row, best_score = r, score
    return best_row


# ---------- команда ----------

class Command(BaseCommand):
    help = "Импорт Excel → Workbook/Sheet + Dataset/DatasetRow (с поддержкой DataTemplate)"

    def add_arguments(self, parser):
        parser.add_argument("path", nargs="?", help="Путь к файлу (если не задан, берётся из --workbook-id)")
        parser.add_argument("--workbook-id", type=int, help="Импорт в существующий Workbook (не создавать новый)")
        parser.add_argument("--sheet", dest="sheet_name", default="", help="Имя листа (по умолчанию первый)")
        parser.add_argument("--header-row", type=int, default=0, help="Номер строки заголовка (1-based). 0 — авто")
        parser.add_argument("--bulk-size", type=int, default=5000, help="Размер пачки для bulk_create DatasetRow")
        parser.add_argument("--dry-run", action="store_true", help="Только разобрать, без записи в БД")
        parser.add_argument("--period-date", dest="period_date", help="Бизнес-дата периода (YYYY-MM-DD)")
        # шаблоны
        parser.add_argument("--template", default="", help="Имя или ID DataTemplate")
        parser.add_argument("--no-auto-template", action="store_true", help="Отключить авто-детект шаблона")

    def handle(self, *args, **opts):
        path = opts.get("path") or ""
        wb_id = opts.get("workbook_id")
        sheet_name = opts.get("sheet_name") or ""
        header_row = int(opts.get("header_row") or 0)
        bulk_size = int(opts.get("bulk_size") or 5000)
        dry = bool(opts.get("dry_run"))

        period_dt = None
        if opts.get("period_date"):
            try:
                period_dt = dtparser.parse(opts["period_date"]).date()
            except Exception:
                raise CommandError(f"Неверный формат --period-date: {opts['period_date']}")

        # 1) Найти реальный путь к файлу
        wb_obj: Optional[Workbook] = None
        if wb_id:
            wb_obj = Workbook.objects.filter(pk=wb_id).first()
            if not wb_obj:
                raise CommandError(f"Workbook id={wb_id} не найден")
            if hasattr(wb_obj, "file") and wb_obj.file:
                path = wb_obj.file.path
            elif path:
                pass
            else:
                raise CommandError("Не удалось определить путь к файлу из Workbook (.file.path)")

        if not path:
            raise CommandError("Не указан путь к файлу и не задан --workbook-id")
        if not os.path.exists(path):
            raise CommandError(f"Файл не найден: {path}")

        # 2) SHA и статусы
        sha = sha256_of_file(path)

        created_wb = False
        if wb_obj:
            wb_obj.sha256 = sha
            wb_obj.status = "importing"
            if not dry:
                wb_obj.save(update_fields=["sha256", "status"])
        else:
            from os.path import basename
            wb_obj = Workbook.objects.create(
                filename=basename(path),
                sha256=sha,
                status="importing",
            )
            created_wb = True

        batch = ImportBatch.objects.create(workbook=wb_obj, status="running", meta={"path": path})

        try:
            # 3) Открыть книгу и лист
            xl = load_workbook(filename=path, read_only=True, data_only=True)
            if sheet_name:
                if sheet_name not in xl.sheetnames:
                    raise CommandError(f"Лист '{sheet_name}' не найден. Доступны: {xl.sheetnames}")
                ws = xl[sheet_name]
            else:
                ws = xl[xl.sheetnames[0]]

            # 4) Заголовки
            if header_row and header_row > ws.max_row:
                raise CommandError(f"--header-row={header_row} за пределами листа (max_row={ws.max_row})")
            hdr_row = header_row or detect_header_row(ws)

            headers: List[str] = []
            for c in range(1, ws.max_column + 1):
                headers.append(str(ws.cell(row=hdr_row, column=c).value or "").strip())

            # 5) Шаблон/маппинг
            dataset_meta: Dict[str, Any] = {}
            selected_template = None
            mapping: Dict[int, str] = {}
            missing: List[str] = []

            Template = apps.get_model("ingest", "DataTemplate")
            tmpl_arg = norm(opts.get("template"))
            auto = not bool(opts.get("no_auto_template"))

            if tmpl_arg:
                try:
                    if tmpl_arg.isdigit():
                        selected_template = Template.objects.prefetch_related("mappings").get(pk=int(tmpl_arg))
                    else:
                        selected_template = Template.objects.prefetch_related("mappings").get(name=tmpl_arg)
                except Template.DoesNotExist:
                    raise CommandError(f"Шаблон '{tmpl_arg}' не найден")
                mapping, missing, _ = xt.match_headers(headers, selected_template)
            elif auto:
                cands = Template.objects.prefetch_related("mappings").all()
                best = xt.detect_best_template(headers, cands)
                if best:
                    selected_template, mapping, missing = best  # type: ignore

            dtype_by_key: Dict[str, str] = {}
            if selected_template:
                if missing:
                    self.stdout.write(self.style.WARNING(
                        f"Шаблон '{selected_template.name}': отсутствуют обязательные поля: {missing}"
                    ))
                dtype_by_key = {m.canonical_key: m.dtype for m in selected_template.mappings.all()}
                dataset_meta.update({
                    "template": selected_template.name,
                    "missing_required": missing,
                    "header_mapping": mapping,  # индекс -> canonical_key
                })

            # 6) Создать Sheet
            sheet_rec = Sheet.objects.create(
                workbook=wb_obj,
                name=ws.title,
                index=xl.sheetnames.index(ws.title),
                n_rows=ws.max_row,
                n_cols=ws.max_column,
            )

            # 7) Создать Dataset (name обязателен)
            base_name = wb_obj.filename or os.path.basename(path)
            dataset_name = f"{base_name} :: {ws.title}"
            dataset = Dataset.objects.create(
                sheet=sheet_rec,
                name=dataset_name,
                inferred_schema={},   # оставляем пустыми (при желании — заполнять)
                primary_key={},
                meta=dataset_meta,
                period_date=period_dt,
                created_at=timezone.now(),
            )

            # 8) Пройти по строкам и bulk_insert DatasetRow
            rows_to_create: List[DatasetRow] = []
            start_row = hdr_row + 1

            for r in range(start_row, ws.max_row + 1):
                row_data: Dict[str, Any] = {}
                empty_row = True

                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c).value
                    if cell not in (None, ""):
                        empty_row = False

                    # ключ: canonical или исходный заголовок
                    if selected_template and (c - 1) in mapping:
                        key = mapping[c - 1]
                        dtype = dtype_by_key.get(key, "text")
                        if dtype == "number":
                            row_data[key] = json_sanitize(coerce_number(cell))
                        elif dtype == "date":
                            row_data[key] = json_sanitize(coerce_date(cell))
                        else:
                            row_data[key] = json_sanitize(cell)
                    else:
                        # fallback — оригинальные заголовки
                        key = headers[c - 1] if c - 1 < len(headers) and headers[c - 1] else f"col_{c}"
                        row_data[key] = json_sanitize(cell)

                if empty_row:
                    continue

                rows_to_create.append(DatasetRow(dataset=dataset, data=row_data))
                if not dry and len(rows_to_create) >= bulk_size:
                    DatasetRow.objects.bulk_create(rows_to_create, batch_size=bulk_size)
                    rows_to_create.clear()

            if not dry and rows_to_create:
                DatasetRow.objects.bulk_create(rows_to_create, batch_size=bulk_size)

            # 9) Финализация
            batch.status = "finished"; batch.save(update_fields=["status"])
            wb_obj.status = "ready";    wb_obj.save(update_fields=["status"])

            self.stdout.write(self.style.SUCCESS(
                f"Imported OK: {os.path.basename(path)} (sheet={ws.title}, rows={max(0, ws.max_row - hdr_row)})"
            ))

        except Exception as e:
            batch.status = "failed"; batch.meta = {"error": str(e)}
            batch.save(update_fields=["status", "meta"])
            wb_obj.status = "error"; wb_obj.save(update_fields=["status"])
            raise
