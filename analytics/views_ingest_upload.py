# analytics/views_ingest_upload.py
from datetime import datetime, date
import json
from decimal import Decimal
from io import BytesIO
from django.db import transaction
from django.utils.timezone import now
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from ingest.models import Workbook, Sheet, Dataset, DatasetRow, DatasetRowRevision, HandleRegistry
from .views_resolve import parse_client_date, format_client_date
from .views_common import user_can_edit_handle
from ingest.models import UploadHistory

try:
    import openpyxl
except Exception:  # нет openpyxl — вернём понятную ошибку при вызове
    openpyxl = None


def _auto_filename(handle: str, period_date):
    # title из HandleRegistry → “Title — DD.MM.YYYY”; если нет — handle
    hr = HandleRegistry.objects.filter(handle=handle).first()
    base = (hr.title or "").strip() if hr and hr.title else handle
    return f"{base} — {format_client_date(period_date)}"


def _get_or_create_workbook(handle: str, period_date, uploaded_by, filename: str | None, sheet_name: str):
    wb = Workbook.objects.filter(handle=handle, period_date=period_date).order_by("-id").first()
    created_wb = False
    if not wb:
        created_wb = True
        wb = Workbook.objects.create(
            filename=filename or _auto_filename(handle, period_date),
            sha256="manual",
            uploaded_by=uploaded_by,
            status=Workbook.STATUS_READY,
            handle=handle,
            period_date=period_date,
            sheets=sheet_name or "Sheet1",
        )
    sh = Sheet.objects.filter(workbook=wb).order_by("index", "id").first()
    if not sh:
        sh = Sheet.objects.create(workbook=wb, name=sheet_name or "Sheet1", index=0)
    ds = Dataset.objects.filter(sheet__workbook=wb).order_by("-created_at", "-id").first()
    created_ds = False
    if not ds:
        created_ds = True
        ds = Dataset.objects.create(
            sheet=sh,
            name=f"{wb.filename} :: {sh.name}",
            meta={"editable": True},
            status=Dataset.STATUS_DRAFT,
            version=1,
            period_date=period_date,
        )
    return wb, sh, ds, created_wb, created_ds


def _grid_from_ws(ws, start_row: int, header_row: int, max_rows: int | None):
    """
    Возвращает 2D-массив (строки x колонки) из openpyxl.Worksheet.
    """
    rows = []
    # openpyxl ws.iter_rows(): 1-based индексация
    # достанем до max_rows строк для экономии
    end_row = ws.max_row if not max_rows else min(ws.max_row, max(header_row, start_row) + max_rows - 1)
    for r in ws.iter_rows(min_row=1, max_row=end_row, values_only=True):
        rows.append(list(r))
    return rows


def _parse_to_records(grid, header_row: int, start_row: int):
    """
    Простой парсер: первая строка с заголовками = header_row.
    Сопоставляем заголовки как есть (без DataTemplate), формируем список dict.
    """
    if not grid or len(grid) < header_row:
        return []
    headers = [(grid[header_row - 1][c] if c < len(grid[header_row - 1]) else None) for c in range(len(grid[header_row - 1]))]
    headers = [ (h if h is not None else f"col_{i+1}") for i, h in enumerate(headers) ]

    records = []
    for r in range(max(start_row - 1, header_row), len(grid)):
        row = grid[r]
        # пустую строку пропускаем
        if not any((cell not in (None, "",)) for cell in row):
            continue
        rec = {}
        for c, h in enumerate(headers):
            val = row[c] if c < len(row) else None
            rec[str(h)] = val
        records.append(rec)
    return records

def _normalize_for_json(obj):
    """Рекурсивно привести данные к JSON-совместимым типам."""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, list):
        return [_normalize_for_json(v) for v in obj]
    if isinstance(obj, dict):
        return {k: _normalize_for_json(v) for k, v in obj.items()}
    return obj


class UploadXLSXView(APIView):
    """
    POST /api/ingest/upload-xlsx/
      multipart/form-data:
        file: <xlsx>
        handle: <slug>         (обязательно)
        period_date: DD.MM.YYYY|YYYY-MM-DD (обязательно)
        sheet_name: "Лист1"    (опционально; по умолчанию активный лист)
        header_row: 1          (опц., по умолчанию 1)
        start_row: 2           (опц., по умолчанию 2)
        max_rows: 5000         (опц., чтобы не тащить мегатаблицы)
        filename: "..."        (опц., если хотите своё имя Workbook)
        truncate: 0|1          (опц., если 1 — очистить старые строки прежде чем писать новую)

    Поведение:
      - Право на редактирование проверяется по HandleRegistry.allowed_users.
      - Создаём (или находим) Workbook(handle, period_date) и «последний» Dataset.
      - Парсим лист Excel в список записей (list[dict]) без шаблонов/валидации.
      - Сохраняем одной строкой: DatasetRow.data = { "parsed": <list[dict]>, "meta": {...} }
        (Если нужно сохранять в несколько строк — легко поменяем.)
      - Если truncate=1 — предварительно удаляем старые строки датасета.

    Ответ: { dataset_id, saved_id, count, period_date, filename }
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @transaction.atomic
    def post(self, request):
        if openpyxl is None:
            return Response({"detail": "openpyxl is not installed on server"}, status=500)

        handle = (request.data.get("handle") or "").strip()
        period_date = parse_client_date(request.data.get("period_date"))
        if not handle or not period_date:
            return Response({"detail": "handle and period_date are required"}, status=400)

        # права
        if not user_can_edit_handle(request.user, handle):
            return Response({"detail": "forbidden for this handle"}, status=403)

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file is required (xlsx)"}, status=400)

        sheet_name = (request.data.get("sheet_name") or "").strip()
        try:
            header_row = int(request.data.get("header_row") or 1)
            start_row  = int(request.data.get("start_row")  or 2)
            max_rows   = request.data.get("max_rows")
            if max_rows is not None:
                max_rows = int(max_rows)
        except Exception:
            return Response({"detail": "header_row/start_row/max_rows must be integers"}, status=400)

        truncate = str(request.data.get("truncate") or "0").lower() in ("1","true","yes")
        filename  = (request.data.get("filename") or "").strip()

        # откроем книгу
        try:
            wb_bytes = file.read()
            wb_obj = openpyxl.load_workbook(filename=BytesIO(wb_bytes), data_only=True, read_only=True)
        except Exception as e:
            return Response({"detail": f"failed to read xlsx: {e}"}, status=400)

        # лист
        if sheet_name:
            if sheet_name not in wb_obj.sheetnames:
                return Response({"detail": f"sheet '{sheet_name}' not found. Available: {', '.join(wb_obj.sheetnames)}"}, status=400)
            ws = wb_obj[sheet_name]
        else:
            ws = wb_obj.active
            sheet_name = ws.title

        # парсинг → grid → records
        grid = _grid_from_ws(ws, start_row=start_row, header_row=header_row, max_rows=max_rows)
        records = _parse_to_records(grid, header_row=header_row, start_row=start_row)

        # upsert workbook/dataset
        wb, sh, ds, created_wb, created_ds = _get_or_create_workbook(
            handle=handle,
            period_date=period_date,
            uploaded_by=request.user,
            filename=filename,
            sheet_name=sheet_name
        )

        if truncate:
            DatasetRow.objects.filter(dataset_id=ds.id).delete()

        saved_ids = []
        for rec in _normalize_for_json(records):
            r = DatasetRow.objects.create(dataset_id=ds.id, data=rec)
            DatasetRowRevision.objects.create(
                row=r,
                version=1,
                data_before={},
                data_after=r.data,
                changed_by=request.user
            )
            saved_ids.append(r.id)

        # --- лог в историю ---
        UploadHistory.objects.create(
            user=request.user,
            handle=handle,
            period_date=period_date,
            workbook=wb,
            dataset=ds,
            filename=wb.filename,
            rows_count=len(saved_ids),
            action=UploadHistory.ACTION_TRUNCATE_UPLOAD if truncate else UploadHistory.ACTION_UPLOAD,
            extra={
                "sheet": sheet_name,
                "header_row": header_row,
                "start_row": start_row,
                "max_rows": max_rows,
            },
        )

        changed = truncate or bool(saved_ids)
        if changed and ds.status == Dataset.STATUS_APPROVED:
            before = ds.status
            ds.status = Dataset.STATUS_DRAFT
            ds.save(update_fields=["status"])
            UploadHistory.objects.create(
                user=request.user,
                handle=handle,
                period_date=period_date,
                workbook=wb,
                dataset=ds,
                filename=wb.filename,
                rows_count=ds.rows.count(),
                action=UploadHistory.ACTION_STATUS_CHANGE,
                status_before=before,
                status_after=ds.status,
                extra={"reason": "auto-draft on data change via upload-xlsx"}
            )

        return Response({
            "dataset_id": ds.id,
            "workbook_id": wb.id,
            "saved_ids": saved_ids,
            "count": len(records),
            "period_date": format_client_date(period_date),
            "filename": wb.filename,
        }, status=200)
